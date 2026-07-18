from __future__ import annotations

import json
import os
import sys
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook

EXPECTED_SHEETS = ["Instructions", "Settings", "Contacts", "Specials", "Lookups"]

SPECIAL_FIELDS = {
    "Sort": "sort",
    "Active": "active",
    "Cut ID": "cutId",
    "Offer Mode": "offerMode",
    "Display Name": "displayName",
    "Brand": "brand",
    "Brand Logo Key": "brandLogoKey",
    "Product Line": "productLine",
    "Marbling Score": "marblingScore",
    "Quantity Available": "quantityAvailable",
    "Primary Price Label": "primaryPriceLabel",
    "Primary Price": "primaryPrice",
    "Primary Image Path": "primaryImagePath",
    "Primary Image Alt": "primaryImageAlt",
    "Primary Image Fit": "primaryImageFit",
    "Primary Image Position": "primaryImagePosition",
    "Secondary Price Label": "secondaryPriceLabel",
    "Secondary Price": "secondaryPrice",
    "Secondary Image Path": "secondaryImagePath",
    "Secondary Image Alt": "secondaryImageAlt",
    "Secondary Image Fit": "secondaryImageFit",
    "Secondary Image Position": "secondaryImagePosition",
    "Savings Message": "savingsMessage",
    "Description": "description",
    "Internal Notes": "internalNotes",
}

CONTACT_FIELDS = {
    "Sort": "sort",
    "Active": "active",
    "Name": "name",
    "Location": "location",
    "Phone": "phone",
    "Email": "email",
}

TRUTHY = {"yes", "true", "1", "active", "show", "visible"}


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "yes" if value else "no"
    return str(value).strip()


def as_bool(value: object) -> bool:
    return clean(value).lower() in TRUTHY


def as_sort(value: object) -> int:
    text = clean(value)
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError as exc:
        raise ValueError(f"Invalid Sort value: {text}") from exc


def table_rows(sheet) -> tuple[list[str], list[list[object]]]:
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    if not rows:
        raise ValueError(f"Sheet {sheet.title} is empty.")

    header_index = next(
        (index for index, row in enumerate(rows) if any(clean(value) for value in row)),
        None,
    )
    if header_index is None:
        raise ValueError(f"Sheet {sheet.title} has no header row.")

    headers = [clean(value) for value in rows[header_index]]
    data_rows = [
        row
        for row in rows[header_index + 1 :]
        if any(clean(value) for value in row)
    ]
    return headers, data_rows


def parse_settings(sheet) -> dict[str, str]:
    headers, rows = table_rows(sheet)
    normalized = {header.lower(): index for index, header in enumerate(headers) if header}
    if "key" not in normalized or "value" not in normalized:
        raise ValueError("Settings sheet must contain Key and Value columns.")

    settings: dict[str, str] = {}
    for row in rows:
        key_index = normalized["key"]
        value_index = normalized["value"]
        key = clean(row[key_index] if key_index < len(row) else "")
        value = clean(row[value_index] if value_index < len(row) else "")
        if key:
            settings[key] = value

    if not settings:
        raise ValueError("Settings sheet contains no settings.")
    return settings


def parse_table(sheet, field_map: dict[str, str]) -> list[dict[str, object]]:
    headers, rows = table_rows(sheet)
    header_map = {header: index for index, header in enumerate(headers) if header}
    missing = [header for header in field_map if header not in header_map]
    if missing:
        raise ValueError(
            f"{sheet.title} sheet is missing required columns: {', '.join(missing)}"
        )

    output: list[dict[str, object]] = []
    for row in rows:
        item: dict[str, object] = {}
        for header, field_name in field_map.items():
            index = header_map[header]
            raw_value = row[index] if index < len(row) else ""
            if field_name == "active":
                item[field_name] = as_bool(raw_value)
            elif field_name == "sort":
                item[field_name] = as_sort(raw_value)
            else:
                item[field_name] = clean(raw_value)
        output.append(item)
    return output


def validate(settings: dict[str, str], contacts: list[dict], specials: list[dict]) -> None:
    active_contacts = [item for item in contacts if item["active"]]
    active_specials = [item for item in specials if item["active"]]

    if not active_contacts:
        raise ValueError("At least one active contact is required.")
    if not active_specials:
        raise ValueError("At least one active special is required.")
    if len(active_specials) > 4:
        raise ValueError("The Legal template supports at most four active specials.")

    allowed_modes = {"dual-offer", "single-offer"}
    allowed_fits = {"cover", "contain"}
    allowed_positions = {
        "center",
        "center top",
        "center bottom",
        "left center",
        "right center",
    }

    seen_cut_ids: set[str] = set()
    for item in active_specials:
        cut_id = str(item["cutId"])
        if not cut_id:
            raise ValueError("Every active special requires Cut ID.")
        if cut_id in seen_cut_ids:
            raise ValueError(f"Duplicate active Cut ID: {cut_id}")
        seen_cut_ids.add(cut_id)

        for field in (
            "displayName",
            "brand",
            "brandLogoKey",
            "productLine",
            "marblingScore",
            "quantityAvailable",
            "primaryPriceLabel",
            "primaryPrice",
            "primaryImagePath",
            "primaryImageAlt",
            "primaryImageFit",
            "primaryImagePosition",
            "savingsMessage",
        ):
            if not clean(item[field]):
                raise ValueError(f"{cut_id} is missing {field}.")

        mode = clean(item["offerMode"])
        if mode not in allowed_modes:
            raise ValueError(f"{cut_id} has invalid Offer Mode: {mode}")

        primary_fit = clean(item["primaryImageFit"])
        primary_position = clean(item["primaryImagePosition"])
        if primary_fit not in allowed_fits:
            raise ValueError(f"{cut_id} has invalid Primary Image Fit: {primary_fit}")
        if primary_position not in allowed_positions:
            raise ValueError(
                f"{cut_id} has invalid Primary Image Position: {primary_position}"
            )

        if mode == "dual-offer":
            for field in (
                "secondaryPriceLabel",
                "secondaryPrice",
                "secondaryImagePath",
                "secondaryImageAlt",
                "secondaryImageFit",
                "secondaryImagePosition",
            ):
                if not clean(item[field]):
                    raise ValueError(f"{cut_id} is dual-offer but is missing {field}.")

            secondary_fit = clean(item["secondaryImageFit"])
            secondary_position = clean(item["secondaryImagePosition"])
            if secondary_fit not in allowed_fits:
                raise ValueError(
                    f"{cut_id} has invalid Secondary Image Fit: {secondary_fit}"
                )
            if secondary_position not in allowed_positions:
                raise ValueError(
                    f"{cut_id} has invalid Secondary Image Position: {secondary_position}"
                )

    footer_url = clean(settings.get("footerUrl", ""))
    if footer_url and not footer_url.lower().startswith("https://"):
        raise ValueError("footerUrl must be blank or start with https://")


def write_json_atomic(output_path: Path, payload: dict) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    serialized = json.dumps(payload, indent=2, ensure_ascii=False) + "\n"
    with NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        newline="\n",
        delete=False,
        dir=output_path.parent,
        prefix=f".{output_path.name}.",
        suffix=".tmp",
    ) as temporary:
        temporary.write(serialized)
        temporary_path = Path(temporary.name)
    os.replace(temporary_path, output_path)


def main() -> int:
    project_root = Path(__file__).resolve().parents[1]
    default_workbook = (
        project_root.parent
        / "Assets"
        / "Live PDF 2.0 Audit"
        / "Paragon Monthly Specials Source V2.xlsx"
    )
    workbook_path = Path(
        os.environ.get("MONTHLY_SPECIALS_V2_WORKBOOK_PATH", str(default_workbook))
    ).expanduser().resolve()
    output_path = project_root / "src" / "data" / "monthly-specials-v2.fixture.json"

    if not workbook_path.is_file():
        raise FileNotFoundError(f"V2 workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    if workbook.sheetnames != EXPECTED_SHEETS:
        raise ValueError(
            "Workbook sheet order mismatch. "
            f"Expected {EXPECTED_SHEETS}, found {workbook.sheetnames}."
        )

    settings = parse_settings(workbook["Settings"])
    contacts = parse_table(workbook["Contacts"], CONTACT_FIELDS)
    specials = parse_table(workbook["Specials"], SPECIAL_FIELDS)
    validate(settings, contacts, specials)
    contacts.sort(key=lambda item: int(item.get("sort", 0)))
    specials.sort(key=lambda item: int(item.get("sort", 0)))

    payload = {
        "source": {
            "type": "workbook",
            "file": str(workbook_path),
        },
        "settings": settings,
        "contacts": contacts,
        "specials": specials,
    }
    write_json_atomic(output_path, payload)

    active_contacts = sum(1 for item in contacts if item["active"])
    active_specials = sum(1 for item in specials if item["active"])
    dual_specials = sum(
        1
        for item in specials
        if item["active"] and item["offerMode"] == "dual-offer"
    )
    single_specials = sum(
        1
        for item in specials
        if item["active"] and item["offerMode"] == "single-offer"
    )

    print("[OK] V2 workbook exported to renderer fixture.")
    print(f"[OK] Workbook: {workbook_path}")
    print(f"[OK] Output: {output_path}")
    print(f"[OK] Settings: {len(settings)}")
    print(f"[OK] Active contacts: {active_contacts}")
    print(
        f"[OK] Active specials: {active_specials} "
        f"({dual_specials} dual, {single_specials} single)"
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(f"[FAIL] {error}", file=sys.stderr)
        raise SystemExit(1)
